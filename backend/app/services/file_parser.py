import csv
import io
from typing import Any

COLUMN_ALIASES = {
    "name": ["name", "productname", "product", "product_name"],
    "sku": ["sku", "productcode", "product_code"],
    "category": ["category", "productcategory", "product_category"],
    "description": ["description"],
    "technical_specs": ["technicalspecs", "technical_specs", "specifications", "specs"],
    "certifications": ["certifications", "certs"],
    "moq": ["moq", "minimumorderquantity", "minimum_order_quantity", "minqty"],
    "pricing": ["pricing"],
    "lead_time_days": ["leadtime", "lead_time", "leadtime_days", "lead_time_days", "deliverydays"],
}


def _normalize(key: str) -> str:
    return key.lower().replace("_", "").replace("-", "").replace(" ", "")


def _parse_number(val: str) -> int | None:
    if not val or not val.strip():
        return None
    cleaned = "".join(c for c in val if c.isdigit() or c in ".-")
    if not cleaned:
        return None
    try:
        return int(float(cleaned))
    except ValueError:
        return None


def _resolve_column(headers: list[str]) -> dict[str, int]:
    header_map: dict[str, str] = {}
    for h in headers:
        header_map[_normalize(h)] = h

    mapping: dict[str, int] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            norm = _normalize(alias)
            if norm in header_map:
                mapping[field] = headers.index(header_map[norm])
                break
    return mapping


def _rows_to_products(rows: list[list[str]]) -> list[dict[str, Any]]:
    if len(rows) < 2:
        raise ValueError("File must have a header row and at least one data row")

    headers = rows[0]
    col_map = _resolve_column(headers)

    if "name" not in col_map:
        raise ValueError("File must have a 'name' column")

    products: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row or all(not cell.strip() for cell in row):
            continue

        def get_val(field: str) -> str:
            idx = col_map.get(field)
            return row[idx].strip() if idx is not None and idx < len(row) else ""

        name = get_val("name")
        if not name:
            continue

        category = get_val("category") or "other"
        category = category.lower().replace(" ", "_")

        products.append({
            "name": name,
            "sku": get_val("sku") or None,
            "category": category,
            "description": get_val("description") or None,
            "technical_specs": get_val("technical_specs") or None,
            "certifications": get_val("certifications") or None,
            "moq": _parse_number(get_val("moq")),
            "unit_price": None,
            "price_range_low": None,
            "price_range_high": None,
            "pricing": get_val("pricing") or None,
            "lead_time_days": _parse_number(get_val("lead_time_days")),
        })

    if not products:
        raise ValueError("No valid product rows found")

    return products


async def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = [[cell.strip() for cell in row] for row in reader]
    return _rows_to_products(rows)


async def _parse_excel(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        rows.append(["" if cell is None else str(cell).strip() for cell in row])
    return _rows_to_products(rows)


async def _parse_docx(content: bytes) -> list[dict[str, Any]]:
    from docx import Document

    doc = Document(io.BytesIO(content))
    rows: list[list[str]] = []
    for table in doc.tables:
        for row in table.rows:
            rows.append([cell.text.strip() for cell in row.cells])
    if not rows:
        raise ValueError("No tables found in Word document; expected a product table")
    return _rows_to_products(rows)


async def _parse_pdf(content: bytes) -> list[dict[str, Any]]:
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(content))
    parts: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        parts.append(text)

    full_text = "\n".join(parts).strip()
    if not full_text:
        raise ValueError("No extractable text found in PDF")

    lines = [line.strip() for line in full_text.splitlines() if line.strip()]
    name = lines[0] if lines else "Imported from PDF"

    return [{
        "name": name,
        "sku": None,
        "category": "other",
        "description": full_text,
        "technical_specs": None,
        "certifications": None,
        "moq": None,
        "unit_price": None,
        "price_range_low": None,
        "price_range_high": None,
        "pricing": None,
        "lead_time_days": None,
    }]


async def parse_file(filename: str, file_content: bytes) -> list[dict[str, Any]]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        return await _parse_csv(file_content)
    if ext == "xlsx":
        return await _parse_excel(file_content)
    if ext == "xls":
        raise ValueError("Legacy .xls format is not supported; please upload .xlsx")
    if ext == "docx":
        return await _parse_docx(file_content)
    if ext == "doc":
        raise ValueError("Legacy .doc format is not supported; please upload .docx")
    if ext == "pdf":
        return await _parse_pdf(file_content)

    raise ValueError(f"Unsupported file type: {ext}")
